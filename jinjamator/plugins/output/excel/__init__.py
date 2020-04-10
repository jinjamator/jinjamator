# Copyright 2019 Wilhelm Putz

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import sys
import os
from jinjamator.tools.output_plugin_base import outputPluginBase
import os
import json
import errno
import json
import xmltodict
from collections import OrderedDict
from flatten_json import flatten
from natsort import natsorted
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


class excel(outputPluginBase):
    def addArguments(self):
        self._parent._parser.add_argument(
            "--excel-output-file-prefix",
            dest="output_filename_prefix",
            help="destination file name prefix [default: %(default)s]",
            default="",
        )
        self._parent._parser.add_argument(
            "--excel-output-file-suffix",
            dest="output_filename_suffix",
            help="destination file name suffix [default: %(default)s]",
            default="xlsx",
        )
        self._parent._parser.add_argument(
            "--excel-output-directory",
            dest="output_directory",
            help="destination directory for generated files",
            default="./",
        )
        self._parent._parser.add_argument(
            "--excel-sheet-name",
            dest="sheet_name",
            help="set excel sheet name[default: %(default)s]",
            default="sheet 1",
        )
        self._parent._parser.add_argument(
            "--excel-order-header",
            dest="order_header",
            help="lexically order header fields [default: %(default)s]",
            default=False,
            action="store_false",
        )
        self._parent._parser.add_argument(
            "--excel-freeze-pane",
            dest="freeze_pane_cell",
            help="set excel freeze pane to cell [default: %(default)s]",
            metavar="<Excel cell name>",
            default="B2",
        )

        self._parent._parser.add_argument(
            "--excel-rename-column",
            dest="rename_columns",
            default=[],
            action="append",
            help="Rename a data colum. Format: <original_name>:<new_name> [default: %(default)s]",
        )

        self._parent._parser.add_argument(
            "--excel-set-columns",
            dest="columns",
            help="Set columns. Format: <col_name_1>:<col_name_2>:<col_name_n> [default: %(default)s]",
        )

    def connect(self, **kwargs):
        pass

    def optimize_column_widths(self, ws):
        column_widths = []
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                try:
                    length = len(cell.value)
                except TypeError:
                    length = 0
                if len(column_widths) > i:
                    if length > column_widths[i]:
                        column_widths[i] = length
                else:
                    column_widths += [length]

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 2

    def get_workbook(self, outputPath, append=False, overwrite=False):
        if os.path.isfile(outputPath) and append:
            self._log.info("appending to %s" % outputPath)
            return load_workbook(outputPath)
        if os.path.isfile(outputPath) and overwrite:
            self._log.info("overwriting %s" % outputPath)
            return Workbook()
        if not os.path.isfile(outputPath):
            self._log.info("creating new file %s" % outputPath)
            return Workbook()
        else:
            self._log.error(
                "destination path %s already exists and neither append or overwrite specified"
                % outputPath
            )
            exit(2)

    def results_to_worksheet(self, data, header, wb, sheet_name, freeze, mergeMode):
        fill = PatternFill(fill_type="solid", start_color="FF000000")
        font = Font(color="FFFFFFFF", bold=True)
        border = Border(
            bottom=Side(border_style="thin", color="FF000000"),
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
        )

        if mergeMode:
            try:
                ws = wb[sheet_name[:30]]
            except BaseException:
                ws = wb.create_sheet(title=sheet_name[:30], index=0)
                ws.append(header)
                for i in range(1, len(header) + 1):
                    ws.cell(row=1, column=i).fill = fill
                    ws.cell(row=1, column=i).font = font
        else:
            ws = wb.create_sheet(title=sheet_name[:30], index=0)
            ws.append(header)
            for i in range(1, len(header) + 1):
                ws.cell(row=1, column=i).fill = fill
                ws.cell(row=1, column=i).font = font
        row = ws.max_row + 1

        for line in data:
            col = 1
            if row % 2 == 0:
                fill = PatternFill(fill_type="solid", start_color="FFCFCFCF")
            else:
                fill = PatternFill(fill_type="solid", start_color="FFEDEDED")
            for item in line.values():
                try:
                    txt = float(item)
                except ValueError:
                    txt = item
                cell = ws.cell(row=row, column=col, value=txt)
                cell.fill = fill
                cell.border = border
                col = col + 1
            row = row + 1

        ws.freeze_panes = freeze
        self.optimize_column_widths(ws)
        ws.auto_filter.ref = ws.dimensions

        return ws

    def process(self, data, **kwargs):
        try:
            os.mkdir(self._parent.configuration._data["output_directory"])
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise
            pass
        dest = "{0}/{1}{2}.{3}".format(
            self._parent.configuration._data["output_directory"],
            self._parent.configuration._data["output_filename_prefix"],
            kwargs["template_path"].replace(os.path.sep, "_"),
            self._parent.configuration._data["output_filename_suffix"],
        )

        if isinstance(data, str):
            try:
                data = json.loads(data)
            except ValueError:
                pass
            try:
                data = xmltodict.parse(data)
            except ValueError:
                pass

        if isinstance(data, dict):
            wb_data = flatten(data)
            keys = list(wb_data.keys())
        elif isinstance(data, list):
            wb_data = []
            for line in data:
                wb_data.append(flatten(line))
            keys = list(wb_data[0].keys())
        if self._parent.configuration.get("order_header"):
            keys = natsorted(keys)

        data = []
        if self._parent.configuration.get("columns"):

            for row in wb_data:
                new_row = OrderedDict()
                for column in self._parent.configuration.get("columns").split(":"):
                    try:
                        new_row[column] = row[column]
                    except KeyError:
                        new_row[column] = ""

                data.append(new_row)
            keys = self._parent.configuration.get("columns").split(":")
        else:
            data = wb_data

        for rename in self._parent.configuration.get("rename_columns"):
            tmp = rename.split(":")
            old = tmp.pop(0)
            new = ":".join(tmp)
            for i, val in enumerate(keys):
                if val == old:
                    keys[i] = new
        wb = Workbook()
        self.results_to_worksheet(
            data,
            keys,
            wb,
            "results",
            self._parent.configuration._data["freeze_pane_cell"],
            False,
        )
        wb.save(dest)

        self._log.info("sucessfully written file: {0}".format(dest))
