# Copyright 2019 Wilhelm Putz

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import openpyxl
import logging
from pprint import pformat
import xxhash
import json
import glob
import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
import xmltodict
from flatten_json import flatten
from natsort import natsorted
from collections import OrderedDict
from copy import deepcopy

__version__ = "0.4"
__author__ = "Wilhelm Putz"

import openpyxl
import logging
import json
import os
import datetime
import xxhash
from openpyxl import load_workbook


class XLSXReader(object):

    def __init__(self, path, worksheet_name, cache=True):
        self.should_cache = cache
        self._log = logging.getLogger("")
        self._path = path
        self._cache = None

        # --------------------------------------------------
        # CACHE
        # --------------------------------------------------
        if cache:
            self.calcCacheFileName(path, worksheet_name)
            try:
                with open(self._cacheFilePath, "r") as fh:
                    self.data = json.loads(fh.read())
                    self._log.info("cache loaded")
                    self._cache = True
                    return
            except Exception:
                self._cache = None

        self.wb = load_workbook(path, data_only=True, read_only=True)

        try:
            self.ws = self.wb[worksheet_name]
        except KeyError:
            self.ws = self.wb[self.wb.sheetnames[0]]

        wb_meta = load_workbook(path, data_only=True, read_only=False)
        self.ws_meta = wb_meta[self.ws.title]

        self._build_merged_map()

    def _build_merged_map(self):
        self.merged_map = {}

        for merged_range in self.ws_meta.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = self.ws_meta.cell(min_row, min_col).value

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    self.merged_map[(r, c)] = value

    def get_value(self, cell, row, col):
        key = (row, col)

        if key in self.merged_map:
            return self.merged_map[key] or ""

        value = getattr(cell, "value", None)
        return value or ""

    def parse_header(self, header_lines=1):
        if self._cache:
            return

        self._dataOffset = header_lines
        self.header = []

        for row in self.ws.iter_rows(min_row=1, max_row=header_lines):
            for col_index, cell in enumerate(row):
                val = str(self.get_value(cell, row=1, col=col_index + 1))

                if len(self.header) <= col_index:
                    self.header.append(val)
                else:
                    self.header[col_index] += " " + val

        self.header = [h.strip() for h in self.header]

    def get_header_for_col(self, col):
        return self.header[col - 1]

    def _iter_data_rows(self,max_empty_streak=50,hard_limit=1_000_000):
        empty_streak = 0
        
        for row_index, row in enumerate(
            self.ws.iter_rows(min_row=1 + self._dataOffset),
            start=1 + self._dataOffset
        ):
            if row_index > hard_limit:
                break

            # fast empty check
            if not any(getattr(cell, "value", None) is not None for cell in row):
                empty_streak += 1
                if empty_streak >= max_empty_streak:
                    break
                continue

            empty_streak = 0
            yield row_index, row

    def parse_data(self):
        if self._cache:
            return

        self._log.info(f"parsing {self._path}")
        self.data = {}

        for row_index, row in self._iter_data_rows():
            assoc = {}

            for col_index, cell in enumerate(row, start=1):
                assoc[self.get_header_for_col(col_index)] = \
                    self.get_value(cell, row_index, col_index)

            self.data[row_index] = assoc

        if self.should_cache:
            with open(self._cacheFilePath, "w") as fh:
                fh.write(json.dumps(self.data, default=self.datetime_handler))

    def datetime_handler(self, x):
        if isinstance(x, datetime.datetime):
            return x.isoformat()
        raise TypeError("Unknown type")

    def calcCacheFileName(self, path, worksheet_name):
        with open(path, "rb") as fh:
            digest = xxhash.xxh64(fh.read()).hexdigest()

        self._cacheFilePath = f"{path}_{worksheet_name}_{digest}.xlsxcache"

    def get_worksheets(self):
        return self.wb.sheetnames

class XLSXWriter(object):
    def __init__(self, path, **kwargs):
        self._log = logging.getLogger()
        self._data = None
        self._order_header = False
        self._header = []
        self._append_sheet = kwargs.get("append_sheet") or False
        self._overwrite_file = kwargs.get("overwrite") or False
        self._destintaion_path = path
        self._freeze_panes = kwargs.get("freeze_pane_cell") or "B1"
        self._column_order = kwargs.get("column_order") or False
        self._rename_columns = kwargs.get("rename_columns") or []
        self._table_style_name = "TableStyleMedium15"
        self._table_name = "Table1"
        self._text_wrap = kwargs.get("text_wrap") or True

        if os.path.isfile(self._destintaion_path) and self._append_sheet:
            self._log.info(f"appending to {self._destintaion_path}")
            self._wb = load_workbook(self._destintaion_path)
        elif os.path.isfile(self._destintaion_path) and self._overwrite_file:
            self._log.info(f"overwriting {self._destintaion_path}")
            self._wb = Workbook()
        elif not os.path.isfile(self._destintaion_path):
            self._log.info(f"creating {self._destintaion_path}")
            self._wb = Workbook()
        else:
            raise ValueError(
                f"destination path {self._destintaion_path} already exists and neither append, nor overwrite specified"
            )

    def sanitize_data(self):

        
        if isinstance(self._data, str):
            try:
                self._data = json.loads(self._data)
            except ValueError:
                pass
            except TypeError:
                pass
            try:
                self._data = xmltodict.parse(self._data)
            except ValueError:
                pass
            except TypeError:
                pass

        if isinstance(self._data, dict):
            self._header = list(self._data.keys())
            is_pseudo_list = True
            for field in self._header:
                try:
                    int(field)
                except ValueError:
                    is_pseudo_list = False
            if is_pseudo_list:  # convert to real list
                
                self._data = list(self._data.values())
            else:
                self._data = flatten(self._data)
                self._header = list(self._data.keys())

        if isinstance(self._data, list):
            data_type = "unknown"
            for index, line in enumerate(self._data):
                data_type = "list_of_dict"
                if isinstance(line, list):
                    data_type = "list_of_list"
                    pass
                elif isinstance(line, dict):
                    self._data[index] = flatten(line)
                else:
                    raise ValueError(
                        f"Line {index} is neither, list nor dict. I don't know how to proceed further {line}"
                    )
                
                #Tried to fix, needs checking
                if data_type == "list_of_dict":
                    self._header = list(self._data[0].keys())
                else:
                    self._header = [
                        f"column {counter}"
                        for counter in range(1, len(self._data[index]) + 1)
                    ]

        if self._column_order:
            data=[]
            for row in self._data:
                new_row = OrderedDict()
                for column in self._column_order:
                    try:
                        new_row[column] = row[column]
                    except KeyError:
                        new_row[column] = ""

                data.append(new_row)
            self._data=data
            self._header = self._column_order

        elif self._order_header:
            self._header = natsorted(self._header)

    def optimize_column_widths(self, ws):
        column_widths = []

        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                if self._text_wrap:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrapText=True
                    )
                try:
                    if (
                        self._text_wrap
                        and isinstance(cell.value, str)
                        and "\n" in cell.value
                    ):
                        length = 0
                        tmp = cell.value.split("\n")
                        for j in tmp:
                            l = len(j)
                            if l > length:
                                length = l
                    else:
                        length = len(cell.value)
                except TypeError:
                    length = 0
                if len(column_widths) > i:
                    if length > column_widths[i]:
                        column_widths[i] = length
                else:
                    column_widths += [length]

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 4

    def save(self):
        self._wb.save(self._destintaion_path)

    def create_sheet_from_data(self, data, sheet_name):
        self._data = data
        self.sanitize_data()
        data = self._data
        header=deepcopy(self._header)
        for old, new in self._rename_columns:
            for i, val in enumerate(self._header):
                if val == old:
                    header[i] = new

        if self._append_sheet:
            try:
                ws = self._wb[sheet_name[:30]]

            except BaseException:
                ws = self._wb.create_sheet(title=sheet_name[:30], index=0)
                ws.append(header)
        else:
            ws = self._wb.create_sheet(title=sheet_name[:30], index=0)
            ws.append(header)
        from pprint import pprint
        
        
        for row in data:
            if isinstance(row,dict):
                values = (row.get(k, "") for k in self._header)
            else:
                values = row
            ws.append(values)

    
        tab = Table(
            displayName=sheet_name[:30].replace(" ", "_"),
            ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}",
        )

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name=self._table_style_name,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style

        ws.add_table(tab)

        ws.freeze_panes = self._freeze_panes
        self.optimize_column_widths(ws)

        return ws


if __name__ == "__main__":
    pass
