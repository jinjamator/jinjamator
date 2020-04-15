# Copyright 2019 Wilhelm Putz

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import openpyxl
import logging
from pprint import pprint, pformat
import xxhash
import json
import glob
import os
import datetime


__version__ = "0.3"
__author__ = "Wilhelm Putz"


class XLSXtoDict(object):
    """
    returns a dict from an xlsx sheet. Respects ranges. Warning currently slow as hell. ca. 300ms/row
    """

    def __init__(self, path, worksheet_name, cache=True):
        self.should_cache = cache
        self._log = logging.getLogger("")
        self._path = path
        self._cache = None
        if cache:
            self.calcCacheFileName(path, worksheet_name)
            try:
                staleCacheFiles = glob.glob(
                    "{0}_{1}_*.xlsxcache".format(path, worksheet_name)
                )
                staleCacheFiles.remove(self._cacheFilePath)
                for staleCacheFile in staleCacheFiles:
                    os.remove(staleCacheFile)
                    self._log.info(
                        "removed stale cache file {0}".format(staleCacheFile)
                    )
            except BaseException:
                pass

            try:
                with open(self._cacheFilePath, "r+b") as fh:
                    self.data = json.loads(fh.read())
                    self._log.info("cache for {0} loaded".format(path))
                    self._cache = True

            except BaseException:
                self._cache = None
                self._log.info("no cache for {0} found".format(path))
        if not self._cache:
            self.wb = openpyxl.load_workbook(path, data_only=True)
            try:
                self.ws = self.wb[worksheet_name]
            except KeyError:
                self._log.warn(
                    "sheet {0} not found, using {1}".format(
                        worksheet_name, self.wb.sheetnames[0]
                    )
                )
                self.ws = self.wb[self.wb.sheetnames[0]]
            self.lastRow = self.get_last_row_in_col()

    def datetime_handler(self, x):
        if isinstance(x, datetime.datetime):
            return x.isoformat()
        raise TypeError("Unknown type")

    def calcCacheFileName(self, path, worksheet_name):
        with open(path, "r+b") as fh:
            digest = xxhash.xxh64(fh.read()).hexdigest()
        self._log.debug("got digest:{0}".format(digest))
        self._cacheFilePath = "{0}_{1}_{2}.xlsxcache".format(
            path, worksheet_name, digest
        )
        self._log.debug("got cache filename:{0}".format(digest))

    def parse_header(self, header_lines=1):
        if self._cache:
            return None
        self.header_fields = []
        if header_lines > 0:
            self._dataOffset = header_lines
        else:
            self._dataOffset = 0
        self._log.debug("detected data offset: {0}".format(self._dataOffset))

        self.merged_cells = {}
        for _range in self.ws.merged_cell_ranges:
            self.merged_cells[str(_range)] = list(
                openpyxl.utils.rows_from_range(str(_range))
            )
        self._log.debug(
            "created merged cell ranges lookup table: {0}".format(self._dataOffset)
        )

        for row_index in range(1, header_lines + 1):
            for cell in self.ws[row_index]:
                try:
                    self.header_fields[cell.column - 1].append(
                        self.get_value_with_merge_lookup(cell)
                    )
                except IndexError:
                    self.header_fields.append([])
                    self.header_fields[cell.column - 1].append(
                        self.get_value_with_merge_lookup(cell)
                    )
        self.header = [" ".join(f).strip() for f in self.header_fields]
        self._log.debug("finished parsing of header")

    def get_last_row_in_col(self, col="A"):
        lastRow = self.ws.max_row

        while self.ws.cell(column=1, row=lastRow).value is None and lastRow > 0:
            lastRow -= 1
        self._log.debug("detected lastRowIn: {0}".format(lastRow))
        return lastRow

    def get_header_for_col(self, col):
        return self.header[col - 1]

    def parse_data(self):
        if self._cache:
            return None
        self._log.info("parsing file: {0}".format(self._path))
        self.data = {}

        for row_index in range(1 + self._dataOffset, self.lastRow + 1):
            self._log.debug("processing row {0}".format(row_index))
            assocRow = {}
            for cell in self.ws[row_index]:
                assocRow[
                    self.get_header_for_col(cell.column)
                ] = self.get_value_with_merge_lookup(cell)
            self.data[row_index] = assocRow
        self._log.info("finished parsing file: {0}".format(self._path))
        self._log.debug(pformat(self.data))
        if self.should_cache:
            self._log.info(
                "saving cachefile to speedup next run: {0}".format(self._cacheFilePath)
            )
            with open(self._cacheFilePath, "w") as fh:
                fh.write(json.dumps(self.data, default=self.datetime_handler))

    def get_value_with_merge_lookup(self, cell):
        for _range, merged_cells in self.merged_cells.items():
            for row in merged_cells:
                if cell.coordinate in row:
                    if not self.ws[merged_cells[0][0]].value:
                        return ""
                    return self.ws[merged_cells[0][0]].value
        if not cell.value:
            return ""
        return cell.value


if __name__ == "__main__":
    pass
